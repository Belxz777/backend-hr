import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.core.cache import cache
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.utils import timezone
import hashlib
import logging
from ...models import Reports, Employee, Functions, Department

logger = logging.getLogger(__name__)

class ReportsExcelExportView(APIView):
    def _generate_cache_key(self, request):
        """Генерация уникального ключа кэша на основе параметров запроса"""
        params = {
            'employee_id': request.query_params.get("employee_id"),
            'function_id': request.query_params.get("function_id"),
            'department_id': request.query_params.get("department_id"),
            'start_date': request.query_params.get("start_date"),
            'end_date': request.query_params.get("end_date"),
        }
        param_string = "&".join(f"{k}={v}" for k, v in sorted(params.items()) if v)
        return f"reports_excel_export:{hashlib.md5(param_string.encode()).hexdigest()}"

    def _format_hours(self, decimal_hours):
        """
        Преобразует десятичные часы в формат ЧЧ:ММ
        2.83 -> 2:50, 1.5 -> 1:30, 0.25 -> 0:15
        """
        try:
            hours = int(decimal_hours)
            minutes_decimal = (decimal_hours - hours) * 60
            minutes = round(minutes_decimal)
            
            # Корректировка если минуты >= 60
            if minutes >= 60:
                hours += 1
                minutes = 0
            
            return f"{hours} ч.:{minutes:02d} мин."
        except (ValueError, TypeError):
            return str(decimal_hours)

    def _calculate_total_hours(self, reports):
        """Вычисляет общее количество часов и возвращает в обоих форматах"""
        total_decimal = sum(float(report.hours_worked) for report in reports)
        total_formatted = self._format_hours(total_decimal)
        return total_decimal, total_formatted

    def get(self, request):
        """Выгрузка отчетов в Excel формате"""
        cache_key = self._generate_cache_key(request)
        cached_response = cache.get(cache_key)
        
        if cached_response is not None:
            logger.debug("Возвращаем кэшированные данные Excel")
            response = HttpResponse(cached_response['content'], content_type=cached_response['content_type'])
            response['Content-Disposition'] = cached_response['content_disposition']
            return response

        try:
            # Параметры фильтрации
            employee_id = request.query_params.get("employee_id")
            function_id = request.query_params.get("function_id")
            department_id = request.query_params.get("department_id")
            start_date_str = request.query_params.get("start_date")
            end_date_str = request.query_params.get("end_date")

            # Фильтрация отчетов
            queryset = Reports.objects.select_related(
                'by_employee', 
                'function',
                'by_employee__department'
            ).all()
            
            if employee_id:
                queryset = queryset.filter(by_employee_id=employee_id)
            
            if function_id:
                queryset = queryset.filter(function_id=function_id)
            
            if department_id:
                queryset = queryset.filter(by_employee__department_id=department_id)
            
            if start_date_str:
                start_date = timezone.datetime.strptime(start_date_str, '%Y-%m-%d').date()
                queryset = queryset.filter(date__date__gte=start_date)
            
            if end_date_str:
                end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d').date()
                queryset = queryset.filter(date__date__lte=end_date)

            # Получение информации о отделе для названия файла
            department_info = ""
            if department_id:
                try:
                    department = Department.objects.get(id=department_id)
                    department_info = f"_department_{department.id}"
                except Department.DoesNotExist:
                    pass

            # Создание Excel файла
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Отчеты"

            # Стили
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            normal_font = Font(size=11)
            normal_alignment = Alignment(vertical="center", wrap_text=True)
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Заголовки столбцов
            headers = [
                "ID отчета", 
                "Код сотрудника", 
                "Сотрудник (ФИО)", 
                "Функция", 
                "Отработанные часы", 
                "Комментарий", 
                "Дата отчета"
            ]
            
            column_widths = [10, 15, 25, 25, 15, 40, 20]
            
            for col_num, (header, width) in enumerate(zip(headers, column_widths), 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                ws.column_dimensions[get_column_letter(col_num)].width = width

            # Заполнение данных с преобразованием часов
            for row_num, report in enumerate(queryset, 2):
                # Полное ФИО сотрудника
                employee_full_name = f"{report.by_employee.surname or ''} {report.by_employee.name} {report.by_employee.patronymic or ''}".strip()
                
                # Преобразование часов в нормальный формат
                formatted_hours = self._format_hours(float(report.hours_worked))
                
                ws.cell(row=row_num, column=1, value=report.id).font = normal_font
                ws.cell(row=row_num, column=2, value=report.by_employee.code).font = normal_font
                ws.cell(row=row_num, column=3, value=employee_full_name).font = normal_font
                ws.cell(row=row_num, column=4, value=report.function.name).font = normal_font
                ws.cell(row=row_num, column=5, value=formatted_hours).font = normal_font
                ws.cell(row=row_num, column=6, value=report.comment or "").font = normal_font
                ws.cell(row=row_num, column=7, value=report.date.strftime('%Y-%m-%d %H:%M')).font = normal_font
                
                # Применение стилей ко всем ячейкам строки
                for col_num in range(1, 8):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.alignment = normal_alignment
                    cell.border = border

            # Добавление итоговой строки
            if queryset.exists():
                total_row = len(queryset) + 2
                total_decimal, total_formatted = self._calculate_total_hours(queryset)
                
                ws.merge_cells(f'A{total_row}:D{total_row}')
                ws.cell(row=total_row, column=1, value="ИТОГО отработанных часов:").font = Font(bold=True)
                ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")
                
                # Отображаем в нормальном формате                
                # Дополнительно можно показать в десятичном формате
                ws.cell(row=total_row, column=5, value=f"({total_decimal:.2f} ч)").font = Font(italic=True, color="808080")
                
                # Стили для итоговой строки
                for col_num in range(1, 8):
                    cell = ws.cell(row=total_row, column=col_num)
                    cell.border = border

            # Добавление информации о фильтрах
            info_row = len(queryset) + 4
            filters_info = []
            
            if start_date_str:
                filters_info.append(f"Начальная дата: {start_date_str}")
            if end_date_str:
                filters_info.append(f"Конечная дата: {end_date_str}")
            if department_id:
                try:
                    department = Department.objects.get(id=department_id)
                    filters_info.append(f"Отдел: {department.name} (ID: {department_id})")
                except Department.DoesNotExist:
                    filters_info.append(f"Отдел ID: {department_id}")
            if employee_id:
                try:
                    employee = Employee.objects.get(id=employee_id)
                    filters_info.append(f"Сотрудник: {employee.name} {employee.surname} (ID: {employee_id})")
                except Employee.DoesNotExist:
                    filters_info.append(f"Сотрудник ID: {employee_id}")
            if function_id:
                try:
                    function = Functions.objects.get(id=function_id)
                    filters_info.append(f"Функция: {function.name} (ID: {function_id})")
                except Functions.DoesNotExist:
                    filters_info.append(f"Функция ID: {function_id}")
            
            if filters_info:
                ws.cell(row=info_row, column=1, value="Параметры фильтрации:").font = Font(bold=True)
                for i, filter_info in enumerate(filters_info, 1):
                    ws.cell(row=info_row + i, column=1, value=filter_info).font = normal_font

            # Сохранение в буфер
            from io import BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # Формирование названия файла с номером отдела и периодом
            period_info = ""
            if start_date_str and end_date_str:
                period_info = f"_{start_date_str}_to_{end_date_str}"
            elif start_date_str:
                period_info = f"_from_{start_date_str}"
            elif end_date_str:
                period_info = f"_to_{end_date_str}"
            
            filename = f"reports{department_info}{period_info}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'

            # Кэширование ответа на 5 минут
            cache_data = {
                'content': buffer.getvalue(),
                'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'content_disposition': f'attachment; filename="{filename}"'
            }
            cache.set(cache_key, cache_data, 300)  # 5 минут

            return response

        except ValueError as e:
            logger.error(f"Ошибка формата даты: {e}")
            return Response(
                {"message": "Неверный формат даты. Используйте YYYY-MM-DD"},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Ошибка при выгрузке Excel: {e}")
            return Response(
                {"message": "Произошла ошибка при выгрузке отчета"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )