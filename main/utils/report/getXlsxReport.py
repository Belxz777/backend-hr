from datetime import datetime, timezone
import io
from django.http import HttpResponseBadRequest, HttpResponseForbidden, StreamingHttpResponse
from openpyxl import Workbook
from requests import Response
from rest_framework.decorators import api_view
from openpyxl.utils import get_column_letter
from main.models import Deputy, Employee,LaborCosts, Functions
from main.serializer import LaborCostsSerializer, FunctionsSerializer
from main.utils.auth import get_user


@api_view(['GET'])
def get_labor_costs_xlsx(request):
    user = get_user(request)

    dep_id = request.query_params.get('id')

    if dep_id:
        labor_costs = LaborCosts.objects.filter(departmentId=user.departmentid.departmentId)
    else:
        labor_costs = LaborCosts.objects.filter(departmentId=dep_id)

    serializer = LaborCostsSerializer(labor_costs, many=True)
    
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Трудозатраты"

    # Заголовки столбцов
    headers = ['report_id', 'employee_id', 'function_id', 'function_name',"is_compulsory", 'spent_time', 'date', 'comment']
    ws.append(headers)
    column_widths = {i: len(header) for i, header in enumerate(headers, start=1)}

    # Заполняем данными
    for cost in serializer.data:
        try:
            if cost['deputyId'] is None:
                func = Functions.objects.get(funcId=cost['functionId'])
                row = [
                    cost['laborCostId'],
                    cost['employeeId'],
                    func.funcId,
                    func.funcName,
                    cost['compulsory'],
                    cost['worked_hours'],
                    cost['date'],
                    cost['comment']
                ]
            else:
                func = Deputy.objects.get(deputyId=cost['deputyId'])
                row = [
                    cost['laborCostId'],
                    cost['employeeId'],
                    cost['deputyId'],
                    func.deputyName,
                    cost['compulsory'],
                    cost['normal_hours'],
                    cost['worked_hours'],
                    cost['date'],
                    cost['comment']
                ]
        
            ws.append(row)
            for i, value in enumerate(row, start=1):
                column_widths[i] = max(column_widths[i], len(str(value)))
        except (Functions.DoesNotExist, Deputy.DoesNotExist):
            continue

    # Устанавливаем ширину столбцов с небольшим запасом
    for i, width in column_widths.items():
        ws.column_dimensions[get_column_letter(i)].width = width + 2

    wb.save(output) 
    output.seek(0)  
    today_date = datetime.now().strftime('%Y-%m-%d')
    filename = f'obshii_otchet_{today_date}.xlsx'

    # Создаем ответ с файлом
    response = StreamingHttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
@api_view(['POST'])
def get_xlsx_precise(request):
    if request.method == 'POST':
        user = get_user(request)
        data = request.data

        employee_ids = data.get('employee_ids', [])
        from_date = data.get('start_date')
        end_date = data.get('end_date')

        if not employee_ids or not from_date or not end_date:
            return HttpResponseBadRequest("Missing required parameters: employee_ids, start_date, or end_date.")

        labor_costs = LaborCosts.objects.filter(
            employeeId__in=employee_ids,
            date__range=[from_date, end_date],
            departmentId=user.departmentid.departmentId
        )

        if not labor_costs.exists():
            output = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Трудозатраты"
            ws.append(["No labor costs found for the given criteria."])
            wb.save(output)
            output.seek(0)

            filename = f'tochnii_otchet_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
            response = StreamingHttpResponse(
                output,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        serializer = LaborCostsSerializer(labor_costs, many=True)

        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Трудозатраты"

        headers = ['report_id', 'employee_id', 'task_id', "tf_name", "avg_time", 'spent_time', 'date', 'comment']
        ws.append(headers)
        column_widths = {i: len(header) for i, header in enumerate(headers, start=1)}

        for cost in serializer.data:
            func = Functions.objects.get(funcId=cost['tf'])
            row = [
                cost['laborCostId'],
                cost['employeeId'],
                func.funcId,
                func.funcName,
                cost['normal_hours'],
                cost['worked_hours'],
                cost['date'],
                cost['comment']
            ]
            ws.append(row)
            for i, value in enumerate(row, start=1):
                column_widths[i] = max(column_widths[i], len(str(value)))

        for i, width in column_widths.items():
            ws.column_dimensions[get_column_letter(i)].width = width + 2

        wb.save(output)
        output.seek(0)

        today_date = datetime.now().strftime('%Y-%m-%d')
        filename = f'tochnii_otchet_{today_date}.xlsx'

        response = StreamingHttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    return HttpResponseBadRequest("Invalid request method. Only POST requests are allowed.")